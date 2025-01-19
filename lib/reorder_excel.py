import pandas as pd
from openpyxl import load_workbook

def process_excel(workbook_name):
    # ====== STEP 1: Initial Data Loading ======
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(workbook_name, sheet_name='Sheet1')
    
    # Store the header from the 4th row (index 3)
    header = df.iloc[2]
    
    # Drop the first 3 rows and last row, keeping data from 5th row onwards
    df = df.iloc[3:-1]
    
    # Set the stored header as column names
    df.columns = header
    
    # ====== STEP 2: Column Setup ======
    # Store original column names for reference
    original_cols = df.columns.tolist()
    
    # Define new columns to be added
    new_columns = ['STOCK_FISICO', 'DIFERENCIA', 'IVA', 'F. VENCIMIENTO', 'SERIAL', 
                   'CODIGO INTERNO', 'MOVILES Y WEB', 'TALLA', 'COLOR']
    
    # Initialize new columns with empty strings
    for col in new_columns:
        df[col] = ''
    
    # ====== STEP 3: Column Reordering ======
    # Define the new column order (numbers represent original column positions)
    new_order = [
        original_cols[0],  # 1st column
        original_cols[2],  # 3rd column
        original_cols[5],  # 6th column
        original_cols[3],  # 4th column
        original_cols[4],  # 5th column
        original_cols[1],  # 2nd column
        original_cols[23], # 24th column
        'STOCK_FISICO',   # new column
        'DIFERENCIA',     # new column
        original_cols[20], # 21st column
        original_cols[21], # 22nd column
        original_cols[18], # 19th column
        'IVA',           # new column
        original_cols[8],  # 9th column
        'F. VENCIMIENTO', # new column
        'SERIAL',        # new column
        'CODIGO INTERNO', # new column
        original_cols[10], # 11th column
        original_cols[12], # 13th column
        original_cols[13], # 14th column
        original_cols[14], # 15th column
        original_cols[15], # 16th column
        original_cols[16], # 17th column
        original_cols[17], # 18th column
        'MOVILES Y WEB',  # new column
        original_cols[6],  # 7th column
        'TALLA',         # new column
        'COLOR'          # new column
    ]
    
    # Apply the new column order
    df = df[new_order]
    
    # ====== STEP 4: Data Cleaning ======
    # Remove apostrophes from string columns
    for column in df.columns:
        if df[column].dtype == 'object':
            df[column] = df[column].astype(str).str.replace("'", "")
    
    # Convert columns 1 and 6 to integers (positions 0 and 2)
    try:
        # Remove non-numeric characters and convert to integers
        df.iloc[:, 0] = pd.to_numeric(df.iloc[:, 0].astype(str).str.replace(r'[^\d.]', ''), errors='coerce').fillna(0).astype(int)
        df.iloc[:, 2] = pd.to_numeric(df.iloc[:, 2].astype(str).str.replace(r'[^\d.]', ''), errors='coerce').fillna(0).astype(int)
    except Exception as e:
        print(f"Warning: Some number conversion errors occurred: {str(e)}")
    
    # ====== STEP 5: Save Initial Changes ======
    # Save the modified DataFrame to Excel
    df.to_excel(workbook_name, sheet_name='Sheet1', index=False)
    
    # ====== STEP 6: Format Excel File ======
    # Load the workbook for formatting
    wb = load_workbook(workbook_name)
    ws = wb['Sheet1']
    
    # Define columns to hide (1-based indexing)
    columns_to_hide = [1, 3, 4, 5] + list(range(12, 29))
    
    # Hide specified columns
    for col_num in columns_to_hide:
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].hidden = True
    
    # Adjust column widths (1-based indexing)
    # Column 2 and 6 made wider
    ws.column_dimensions[ws.cell(row=1, column=2).column_letter].width = 60  # Adjust this value as needed
    ws.column_dimensions[ws.cell(row=1, column=6).column_letter].width = 20  # Adjust this value as needed
    
    # Save all changes
    wb.save(workbook_name)
    
    print("Excel file processed successfully!")

# ====== STEP 7: User Input and Error Handling ======
# Get workbook name from user
workbook_name = input("Please enter the Excel workbook name (including .xlsx extension): ")

try:
    process_excel(workbook_name)
except FileNotFoundError:
    print("Error: File not found. Please check the file name and try again.")
except IndexError:
    print("Error: The file doesn't have enough columns. Please check the file structure.")
except Exception as e:
    print(f"An error occurred: {str(e)}")
